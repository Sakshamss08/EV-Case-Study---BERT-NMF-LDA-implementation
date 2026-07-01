import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

# =====================================================================
# CONFIGURATION & FILE PATHS
# =====================================================================
# Using pathlib makes path handling safer across different operating systems
BASE_DIR = Path("C:/Users/saksham singhal/Desktop/EV/files/data/")
OUTPUT_FILE = BASE_DIR / "EV_Customer_Insights_Topic_Modeling.xlsx"

# =====================================================================
# EXCEL STYLING CONSTANTS
# =====================================================================
FONT_NAME = "Arial"

# Colors & Fills
HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
STRENGTH_FILL = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE") # Light Green
PAIN_FILL = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")     # Light Red

# Typography
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="595959")
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
BODY_FONT = Font(name=FONT_NAME, size=10)

# Borders
THIN_BORDER_STYLE = Side(style="thin", color="D9D9D9")
STANDARD_BORDER = Border(
    left=THIN_BORDER_STYLE, 
    right=THIN_BORDER_STYLE, 
    top=THIN_BORDER_STYLE, 
    bottom=THIN_BORDER_STYLE
)

# =====================================================================
# HELPER FUNCTIONS
# =====================================================================

def style_header_row(worksheet, row_idx, num_columns):
    """Applies standard corporate styling to header rows in the Excel sheet."""
    for col_idx in range(1, num_columns + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = STANDARD_BORDER

def autofit_columns(worksheet, dataframe, start_col=1, min_width=10, max_width=60):
    """Adjusts column widths based on the length of the data to prevent text clipping."""
    for i, column_name in enumerate(dataframe.columns):
        col_letter = get_column_letter(start_col + i)
        
        # Calculate max length looking at the header and the first 200 rows of data
        sample_data = dataframe[column_name].astype(str).values[:200]
        max_len = max([len(str(column_name))] + [len(str(val)) for val in sample_data])
        
        # Constrain width between min and max parameters
        adjusted_width = min(max(max_len + 2, min_width), max_width)
        worksheet.column_dimensions[col_letter].width = adjusted_width

def write_dataframe_to_sheet(worksheet, dataframe, start_row=1, start_col=1, include_header=True):
    """Writes a Pandas DataFrame to an openpyxl worksheet and applies standard styling."""
    current_row = start_row
    
    if include_header:
        for j, col_name in enumerate(dataframe.columns):
            worksheet.cell(row=start_row, column=start_col + j, value=col_name)
        style_header_row(worksheet, start_row, len(dataframe.columns))
        current_row += 1

    for row_tuple in dataframe.itertuples(index=False):
        for j, val in enumerate(row_tuple):
            cell = worksheet.cell(row=current_row, column=start_col + j, value=val)
            cell.font = BODY_FONT
            cell.border = STANDARD_BORDER
        current_row += 1
        
    return current_row - 1  # Return the index of the last row written

# =====================================================================
# MAIN EXECUTION
# =====================================================================

def main():
    print("Loading data files...")
    # Load all datasets
    final_df = pd.read_csv(BASE_DIR / "ev_customer_feedback_final.csv")
    lda_summary = pd.read_csv(BASE_DIR / "lda_topic_summary.csv")
    nmf_summary = pd.read_csv(BASE_DIR / "nmf_topic_summary.csv")
    bert_summary = pd.read_csv(BASE_DIR / "bertopic_topic_summary.csv")
    theme_prevalence = pd.read_csv(BASE_DIR / "theme_prevalence_cross_model.csv")
    theme_by_product = pd.read_csv(BASE_DIR / "theme_by_product_type.csv")
    pain_vs_strength = pd.read_csv(BASE_DIR / "pain_points_vs_strengths.csv")

    wb = Workbook()

    # -----------------------------------------------------------------
    # SHEET 1: Executive Summary
    # -----------------------------------------------------------------
    exec_summary_sheet = wb.active
    exec_summary_sheet.title = "Executive Summary"
    exec_summary_sheet.sheet_view.showGridLines = False

    # Titles and Context
    exec_summary_sheet["B2"] = "EV Car / Scooter Launch | Customer Feedbacks"
    exec_summary_sheet["B2"].font = TITLE_FONT
    exec_summary_sheet["B3"] = "Topic Modeling Analysis of Customer Feedback (LDA, NMF, BERTopic) — Corpus: 1,200 feedback records"
    exec_summary_sheet["B3"].font = SUBTITLE_FONT

    # Section 1: Objective
    exec_summary_sheet["B5"] = "1. Objective"
    exec_summary_sheet["B5"].font = SECTION_FONT
    
    objective_text = (
        "Identify the themes that matter most to prospective EV Car / EV Scooter customers, "
        "using three independent topic-modeling techniques (LDA, NMF, BERTopic) applied to "
        "1,200 pieces of customer feedback (surveys, social media, support tickets, marketplace "
        "reviews) across 12 competitor EV car and scooter brands."
    )
    exec_summary_sheet["B6"] = objective_text
    exec_summary_sheet["B6"].font = BODY_FONT
    exec_summary_sheet["B6"].alignment = Alignment(wrap_text=True)
    exec_summary_sheet.merge_cells("B6:J6")
    exec_summary_sheet.row_dimensions[6].height = 45

    # Section 2: Ranked Themes
    exec_summary_sheet["B8"] = "2. Ranked Themes (Average Prevalence Across All 3 Models)"
    exec_summary_sheet["B8"].font = SECTION_FONT

    # Format dataframe for display
    theme_prevalence_display = theme_prevalence.rename(columns={theme_prevalence.columns[0]: "Business Theme"})
    last_row_themes = write_dataframe_to_sheet(exec_summary_sheet, theme_prevalence_display, start_row=9, start_col=2)
    autofit_columns(exec_summary_sheet, theme_prevalence_display, start_col=2)

    # Insert Bar Chart for Themes
    chart = BarChart()
    chart.title = "What Matters Most to Customers (Avg. % Across 3 Models)"
    chart.y_axis.title = "% of feedback"
    chart.x_axis.title = "Theme"
    
    num_themes = len(theme_prevalence_display)
    chart_data = Reference(exec_summary_sheet, min_col=6, min_row=9, max_row=9 + num_themes, max_col=6)  # Average_pct column
    chart_cats = Reference(exec_summary_sheet, min_col=2, min_row=10, max_row=9 + num_themes)
    
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_cats)
    chart.width = 22
    chart.height = 11
    
    chart_anchor = "B" + str(last_row_themes + 3)
    exec_summary_sheet.add_chart(chart, chart_anchor)

    # Section 3: Pain Points vs Strengths
    row_after_chart = last_row_themes + 25
    exec_summary_sheet.cell(row=row_after_chart, column=2, value="3. Pain Points vs Strengths (avg. rating by theme)").font = SECTION_FONT
    
    pain_vs_strength_display = pain_vs_strength.rename(columns={
        pain_vs_strength.columns[0]: "Business Theme",
        "avg_rating": "Avg Rating (1-5)",
        "doc_count": "Feedback Volume"
    })
    
    last_row_pain_points = write_dataframe_to_sheet(exec_summary_sheet, pain_vs_strength_display, start_row=row_after_chart + 1, start_col=2)
    autofit_columns(exec_summary_sheet, pain_vs_strength_display, start_col=2)

    # Color code pain points (Red) and strengths (Green) based on thresholds
    for r in range(row_after_chart + 2, last_row_pain_points + 1):
        rating_cell = exec_summary_sheet.cell(row=r, column=3)
        try:
            rating_value = float(rating_cell.value)
            if rating_value < 3.9:  # Threshold for pain points
                for c in range(2, 5):
                    exec_summary_sheet.cell(row=r, column=c).fill = PAIN_FILL
            elif rating_value > 4.1:  # Threshold for strengths
                for c in range(2, 5):
                    exec_summary_sheet.cell(row=r, column=c).fill = STRENGTH_FILL
        except (TypeError, ValueError):
            pass

    # Section 4: Key Takeaways
    takeaways_start_row = last_row_pain_points + 3
    exec_summary_sheet.cell(row=takeaways_start_row, column=2, value="4. Key Takeaways for Launch Strategy").font = SECTION_FONT
    
    key_takeaways = [
        "• Charging Infrastructure and Price/Subsidy show the LOWEST average ratings (~3.8) despite high feedback volume — these are the top pain points to solve before/at launch.",
        "• Environmental Impact and Software/Connectivity are the MOST DISCUSSED themes across all 3 models (~19% average prevalence each) — customers are highly vocal on sustainability messaging and app/software experience.",
        "• Performance & Ride Quality and Design/Build Quality score as relative STRENGTHS (avg rating > 4.0) — good differentiators to lead with in marketing.",
        "• Theme prevalence is broadly SIMILAR between EV Car and EV Scooter buyers — Environmental Impact, Software/Connectivity and Performance top both segments — but Scooter buyers raise Charging Infrastructure slightly more often (proximity/frequency of daily charging).",
        "• Cross-model agreement on exact business theme per document is 13.3% (all 3 agree) and 47.1% (at least 2 of 3 agree) — expected given each algorithm's different mathematical approach; convergent themes across models are the most trustworthy signals."
    ]
    
    for i, takeaway in enumerate(key_takeaways):
        current_row = takeaways_start_row + 1 + i
        cell = exec_summary_sheet.cell(row=current_row, column=2, value=takeaway)
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        exec_summary_sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
        exec_summary_sheet.row_dimensions[current_row].height = 30

    # formatting margins
    exec_summary_sheet.column_dimensions["A"].width = 3

    # -----------------------------------------------------------------
    # SHEET 2: Raw Dataset
    # -----------------------------------------------------------------
    raw_data_sheet = wb.create_sheet("Raw Feedback Data")
    raw_cols = ["review_id", "source", "product_type", "brand", "region", "date", "rating", "feedback_text"]
    
    write_dataframe_to_sheet(raw_data_sheet, final_df[raw_cols], start_row=1, start_col=1)
    autofit_columns(raw_data_sheet, final_df[raw_cols])
    raw_data_sheet.column_dimensions["H"].width = 70
    
    # Format as an Excel Table for easy sorting/filtering
    raw_table = Table(displayName="RawFeedback", ref=f"A1:H{len(final_df)+1}")
    raw_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    raw_data_sheet.add_table(raw_table)
    raw_data_sheet.freeze_panes = "A2"

    # -----------------------------------------------------------------
    # SHEET 3: LDA Results
    # -----------------------------------------------------------------
    lda_sheet = wb.create_sheet("LDA Results")
    lda_sheet["A1"] = "LDA (Latent Dirichlet Allocation) — Topic Summary"
    lda_sheet["A1"].font = SECTION_FONT
    lda_sheet.merge_cells("A1:E1")
    
    write_dataframe_to_sheet(lda_sheet, lda_summary, start_row=3, start_col=1)
    autofit_columns(lda_sheet, lda_summary)
    lda_sheet.column_dimensions["B"].width = 60

    doc_cols_lda = ["review_id", "product_type", "brand", "rating", "feedback_text", "lda_topic", "lda_topic_label", "lda_topic_confidence"]
    lda_docs_start = 3 + len(lda_summary) + 3
    lda_sheet.cell(row=lda_docs_start - 1, column=1, value="Document-Level Topic Assignments (sample)").font = SECTION_FONT
    
    write_dataframe_to_sheet(lda_sheet, final_df[doc_cols_lda], start_row=lda_docs_start, start_col=1)
    autofit_columns(lda_sheet, final_df[doc_cols_lda])
    lda_sheet.column_dimensions["E"].width = 70
    lda_sheet.column_dimensions["G"].width = 45

    # -----------------------------------------------------------------
    # SHEET 4: NMF Results
    # -----------------------------------------------------------------
    nmf_sheet = wb.create_sheet("NMF Results")
    nmf_sheet["A1"] = "NMF (Non-negative Matrix Factorization) — Topic Summary"
    nmf_sheet["A1"].font = SECTION_FONT
    nmf_sheet.merge_cells("A1:E1")
    
    write_dataframe_to_sheet(nmf_sheet, nmf_summary, start_row=3, start_col=1)
    autofit_columns(nmf_sheet, nmf_summary)
    nmf_sheet.column_dimensions["B"].width = 60

    doc_cols_nmf = ["review_id", "product_type", "brand", "rating", "feedback_text", "nmf_topic", "nmf_topic_label", "nmf_topic_confidence"]
    nmf_docs_start = 3 + len(nmf_summary) + 3
    nmf_sheet.cell(row=nmf_docs_start - 1, column=1, value="Document-Level Topic Assignments (sample)").font = SECTION_FONT
    
    write_dataframe_to_sheet(nmf_sheet, final_df[doc_cols_nmf], start_row=nmf_docs_start, start_col=1)
    autofit_columns(nmf_sheet, final_df[doc_cols_nmf])
    nmf_sheet.column_dimensions["E"].width = 70
    nmf_sheet.column_dimensions["G"].width = 45

    # -----------------------------------------------------------------
    # SHEET 5: BERTopic Results
    # -----------------------------------------------------------------
    bert_sheet = wb.create_sheet("BERTopic Results")
    bert_sheet["A1"] = "BERTopic (Transformer-based Topic Modeling) — Topic Summary"
    bert_sheet["A1"].font = SECTION_FONT
    bert_sheet.merge_cells("A1:E1")
    
    bert_summary_clean = bert_summary[["Topic", "Count", "Name", "Representation"]].copy()
    bert_summary_clean["Representation"] = bert_summary_clean["Representation"].astype(str)
    
    write_dataframe_to_sheet(bert_sheet, bert_summary_clean, start_row=3, start_col=1)
    autofit_columns(bert_sheet, bert_summary_clean)
    bert_sheet.column_dimensions["C"].width = 45
    bert_sheet.column_dimensions["D"].width = 70

    doc_cols_bert = ["review_id", "product_type", "brand", "rating", "feedback_text", "bert_topic", "bert_topic_label", "bert_topic_confidence"]
    bert_docs_start = 3 + len(bert_summary_clean) + 3
    bert_sheet.cell(row=bert_docs_start - 1, column=1, value="Document-Level Topic Assignments (sample)").font = SECTION_FONT
    
    write_dataframe_to_sheet(bert_sheet, final_df[doc_cols_bert], start_row=bert_docs_start, start_col=1)
    autofit_columns(bert_sheet, final_df[doc_cols_bert])
    bert_sheet.column_dimensions["E"].width = 70
    bert_sheet.column_dimensions["G"].width = 45

    # -----------------------------------------------------------------
    # SHEET 6: Cross-Model Comparison
    # -----------------------------------------------------------------
    compare_sheet = wb.create_sheet("Cross-Model Comparison")
    compare_sheet["A1"] = "Business Theme Mapping — LDA vs NMF vs BERTopic (per document)"
    compare_sheet["A1"].font = SECTION_FONT
    compare_sheet.merge_cells("A1:F1")

    compare_cols = [
        "review_id", "product_type", "feedback_text",
        "lda_business_theme", "nmf_business_theme", "bert_business_theme",
        "models_agree_all3", "models_agree_2of3"
    ]
    
    compare_df = final_df[compare_cols].rename(columns={
        "lda_business_theme": "LDA Theme",
        "nmf_business_theme": "NMF Theme",
        "bert_business_theme": "BERTopic Theme",
        "models_agree_all3": "All 3 Agree?",
        "models_agree_2of3": "At Least 2 Agree?"
    })
    
    write_dataframe_to_sheet(compare_sheet, compare_df, start_row=3, start_col=1)
    autofit_columns(compare_sheet, compare_df)
    compare_sheet.column_dimensions["C"].width = 70
    compare_sheet.freeze_panes = "A4"

    # Add Theme Prevalence to the side
    compare_sheet.cell(row=1, column=8, value="Theme Prevalence by Product Type").font = SECTION_FONT
    theme_by_product_display = theme_by_product.rename(columns={theme_by_product.columns[0]: "Business Theme"})
    write_dataframe_to_sheet(compare_sheet, theme_by_product_display, start_row=3, start_col=9)

    # -----------------------------------------------------------------
    # SAVE EXCEL FILE
    # -----------------------------------------------------------------
    wb.save(OUTPUT_FILE)
    print(f"Excel workbook successfully saved to: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()